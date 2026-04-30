"""Build the consolidated Excel snapshot from the page manifest + structured.json."""
from __future__ import annotations
import json
from collections import defaultdict
from pathlib import Path
import polars as pl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from ..config import settings
from ..models import PageInfo, PageType
from ..ingest.codes import family_hint


HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center")


def _all_code_mentions(infos: list[PageInfo]) -> pl.DataFrame:
    rows = []
    for p in infos:
        for c in p.codes:
            rows.append(
                {
                    "code": c,
                    "page": p.page_num,
                    "section_l1": p.section_l1,
                    "section_l2": p.section_l2,
                    "page_type": p.page_type.value,
                    "family_hint": family_hint(c),
                }
            )
    return pl.DataFrame(rows) if rows else pl.DataFrame(
        schema={
            "code": pl.Utf8,
            "page": pl.Int64,
            "section_l1": pl.Utf8,
            "section_l2": pl.Utf8,
            "page_type": pl.Utf8,
            "family_hint": pl.Utf8,
        }
    )


def _codes_consolidated(mentions: pl.DataFrame) -> pl.DataFrame:
    if mentions.is_empty():
        return mentions
    # First-seen page (canonical), and aggregated metadata.
    grouped = (
        mentions.group_by("code")
        .agg(
            pl.col("page").min().alias("first_page"),
            pl.col("page").n_unique().alias("nb_pages"),
            pl.col("page").alias("pages"),
            pl.col("section_l1").drop_nulls().unique().alias("sections_l1"),
            pl.col("section_l2").drop_nulls().unique().alias("sections_l2"),
            pl.col("page_type").unique().alias("page_types"),
            pl.col("family_hint").drop_nulls().first().alias("family_hint"),
        )
        .sort("code")
        .with_columns(
            pl.col("pages").list.eval(pl.element().cast(pl.Utf8)).list.join(", "),
            pl.col("sections_l1").list.join(" | "),
            pl.col("sections_l2").list.join(" | "),
            pl.col("page_types").list.join(", "),
        )
    )
    return grouped


def _filter_by_pagetype(mentions: pl.DataFrame, types: set[PageType]) -> pl.DataFrame:
    if mentions.is_empty():
        return mentions
    type_values = [t.value for t in types]
    return mentions.filter(pl.col("page_type").is_in(type_values))


def _pages_overview(infos: list[PageInfo]) -> pl.DataFrame:
    return pl.DataFrame(
        [
            {
                "page": p.page_num,
                "section_l1": p.section_l1,
                "section_l2": p.section_l2,
                "page_type": p.page_type.value,
                "n_codes": len(p.codes),
                "text_len": p.text_len,
                "n_drawings": p.n_drawings,
                "n_images": p.n_images,
            }
            for p in infos
        ]
    )


def _stats(infos: list[PageInfo], mentions: pl.DataFrame) -> pl.DataFrame:
    by_type: dict[str, int] = defaultdict(int)
    for p in infos:
        by_type[p.page_type.value] += 1
    rows = [
        {"metric": "Pages totales", "valeur": len(infos)},
        {"metric": "Codes uniques", "valeur": mentions["code"].n_unique() if not mentions.is_empty() else 0},
        {"metric": "Mentions de codes", "valeur": mentions.height},
    ]
    for t, n in sorted(by_type.items()):
        rows.append({"metric": f"Pages [{t}]", "valeur": n})
    return pl.DataFrame(rows)


def _coerce(v):
    if v is None:
        return None
    if isinstance(v, (list, tuple)):
        return ", ".join(str(x) for x in v)
    if isinstance(v, dict):
        import json as _json
        return _json.dumps(v, ensure_ascii=False)
    return v


def _write_sheet(wb: Workbook, name: str, df: pl.DataFrame) -> None:
    ws = wb.create_sheet(title=name[:31])
    if df.is_empty():
        ws.append(["(vide)"])
        return
    headers = df.columns
    ws.append(headers)
    for row in df.iter_rows():
        ws.append([_coerce(v) for v in row])

    # Header style
    for col_idx, _ in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col_idx)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = HEADER_ALIGN

    # Auto column width (rough)
    for col_idx, h in enumerate(headers, start=1):
        max_len = max(
            [len(str(h))]
            + [len(str(v)) if v is not None else 0 for v in df[h].to_list()[:200]]
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)

    # Freeze first row
    ws.freeze_panes = "A2"

    # Excel table
    last_col = get_column_letter(len(headers))
    last_row = df.height + 1
    table = Table(displayName=f"T_{name.replace(' ', '_')[:25]}", ref=f"A1:{last_col}{last_row}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9", showRowStripes=True
    )
    ws.add_table(table)


def _load_structured() -> dict:
    p = settings.output_dir / "structured.json"
    if not p.exists():
        return {"profiles": [], "performance": [], "vitrage": [], "coupes": []}
    return json.loads(p.read_text(encoding="utf-8"))


def _df_or_empty(rows: list[dict], schema: dict[str, type] | None = None) -> pl.DataFrame:
    if rows:
        return pl.DataFrame(rows)
    return pl.DataFrame(schema=schema) if schema else pl.DataFrame()


def build_excel(infos: list[PageInfo], output_path: Path | None = None) -> Path:
    out = output_path or (settings.output_dir / "catalog_perf70_gti_v2.xlsx")
    out.parent.mkdir(parents=True, exist_ok=True)

    mentions = _all_code_mentions(infos)
    consolidated = _codes_consolidated(mentions)
    profiles_mentions = _filter_by_pagetype(mentions, {PageType.NOMENCLATURE})
    plans_debit = _filter_by_pagetype(mentions, {PageType.PLAN_DEBIT})
    coupes_mentions = _filter_by_pagetype(mentions, {PageType.COUPE})
    perf_mentions = _filter_by_pagetype(mentions, {PageType.PERFORMANCE})
    vitrage_mentions = _filter_by_pagetype(mentions, {PageType.VITRAGE})
    montage = _filter_by_pagetype(mentions, {PageType.PLAN_MONTAGE})
    pages = _pages_overview(infos)
    stats = _stats(infos, mentions)

    structured = _load_structured()
    profiles_struct = _df_or_empty(structured.get("profiles") or [])
    perf_struct = _df_or_empty(structured.get("performance") or [])
    vitrage_struct = _df_or_empty(structured.get("vitrage") or [])
    coupes_struct = _df_or_empty(structured.get("coupes") or [])

    wb = Workbook()
    wb.remove(wb.active)

    _write_sheet(wb, "Stats", stats)
    _write_sheet(wb, "Codes (consolidé)", consolidated)
    # Structured (VLM) sheets first — the "ready-to-use" tabs
    _write_sheet(wb, "Profilés (structuré)", profiles_struct)
    _write_sheet(wb, "Performances thermiques", perf_struct)
    _write_sheet(wb, "Tableau vitrage", vitrage_struct)
    _write_sheet(wb, "Coupes (annotées)", coupes_struct)
    # Regex-only sheets (mentions per page)
    _write_sheet(wb, "Profilés (mentions)", profiles_mentions)
    _write_sheet(wb, "Plans de débits", plans_debit)
    _write_sheet(wb, "Coupes (mentions)", coupes_mentions)
    _write_sheet(wb, "Performances (mentions)", perf_mentions)
    _write_sheet(wb, "Vitrage (mentions)", vitrage_mentions)
    _write_sheet(wb, "Plans de montage", montage)
    _write_sheet(wb, "Mentions (toutes)", mentions)
    _write_sheet(wb, "Pages (overview)", pages)

    wb.save(out)
    return out
